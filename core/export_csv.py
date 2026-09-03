"""
HARMATTAN — CSV, XLSX, and Markdown export builders.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any


def build_csv_report(data: dict) -> str:
    """Build a CSV string report from session data."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["HARMATTAN Export", datetime.now().isoformat()])
    writer.writerow([])

    # Hosts
    hosts = data.get("arp", {}).get("hosts", data.get("hosts", []))
    if hosts:
        writer.writerow(["=== HÔTES ==="])
        writer.writerow(["IP", "MAC", "Vendeur", "Hostname", "Rôle", "Première vue", "Dernière vue"])
        for h in hosts:
            writer.writerow([
                h.get("ip", ""), h.get("mac", ""), h.get("vendor", ""),
                h.get("hostname", ""), h.get("role", ""),
                h.get("first_seen", ""), h.get("last_seen", ""),
            ])
        writer.writerow([])

    # Ports / findings
    attack = data.get("attack_surface", {})
    exposures = attack.get("exposures", attack.get("findings", []))
    if exposures:
        writer.writerow(["=== EXPOSITIONS ==="])
        writer.writerow(["Hôte", "Port", "Service", "Risque", "Détail"])
        for exp in exposures:
            writer.writerow([
                exp.get("host", exp.get("ip", "")),
                exp.get("port", ""),
                exp.get("service", ""),
                exp.get("risk", exp.get("severity", "")),
                exp.get("detail", "")[:100],
            ])
        writer.writerow([])

    # CVE
    cves = data.get("vuln", {}).get("cves", data.get("cves", []))
    if cves:
        writer.writerow(["=== CVE ==="])
        writer.writerow(["CVE ID", "Score", "Sévérité", "Description"])
        for cve in cves:
            writer.writerow([
                cve.get("id", cve.get("cve_id", "")),
                cve.get("score", ""),
                cve.get("severity", ""),
                cve.get("description", "")[:120],
            ])

    return output.getvalue()


def build_xlsx_report(data: dict, filepath: str | Path) -> None:
    """Build an Excel workbook from session data."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX export (pip install openpyxl)")

    wb = Workbook()

    # Sheet 1: Hosts
    ws = wb.active
    ws.title = "Hôtes"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    headers = ["IP", "MAC", "Vendeur", "Hostname", "Rôle", "Première vue", "Dernière vue"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    hosts = data.get("arp", {}).get("hosts", data.get("hosts", []))
    for row, h in enumerate(hosts, 2):
        for col, key in enumerate(["ip", "mac", "vendor", "hostname", "role", "first_seen", "last_seen"], 1):
            ws.cell(row=row, column=col, value=h.get(key, ""))

    # Sheet 2: Expositions
    ws2 = wb.create_sheet("Expositions")
    headers2 = ["Hôte", "Port", "Service", "Risque", "Détail"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    attack = data.get("attack_surface", {})
    exposures = attack.get("exposures", attack.get("findings", []))
    for row, exp in enumerate(exposures, 2):
        ws2.cell(row=row, column=1, value=exp.get("host", exp.get("ip", "")))
        ws2.cell(row=row, column=2, value=exp.get("port", ""))
        ws2.cell(row=row, column=3, value=exp.get("service", ""))
        ws2.cell(row=row, column=4, value=exp.get("risk", exp.get("severity", "")))
        ws2.cell(row=row, column=5, value=exp.get("detail", "")[:100])

    # Sheet 3: CVE
    cves = data.get("vuln", {}).get("cves", data.get("cves", []))
    if cves:
        ws3 = wb.create_sheet("CVE")
        headers3 = ["CVE ID", "Score", "Sévérité", "Description"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row, cve in enumerate(cves, 2):
            ws3.cell(row=row, column=1, value=cve.get("id", cve.get("cve_id", "")))
            ws3.cell(row=row, column=2, value=cve.get("score", ""))
            ws3.cell(row=row, column=3, value=cve.get("severity", ""))
            ws3.cell(row=row, column=4, value=cve.get("description", "")[:120])

    wb.save(str(filepath))


def build_markdown_report(data: dict) -> str:
    """Build a Markdown report from session data."""
    lines = []
    lines.append("# HARMATTAN Audit Report")
    lines.append(f"**Date**: {datetime.now().isoformat()}")
    lines.append("")

    # Summary
    hosts = data.get("arp", {}).get("hosts", data.get("hosts", []))
    lines.append("## Résumé")
    lines.append(f"- **Hôtes découverts**: {len(hosts)}")
    attack = data.get("attack_surface", {})
    exposures = attack.get("exposures", attack.get("findings", []))
    lines.append(f"- **Expositions**: {len(exposures)}")
    cves = data.get("vuln", {}).get("cves", data.get("cves", []))
    lines.append(f"- **CVE**: {len(cves)}")
    lines.append("")

    # Hosts table
    if hosts:
        lines.append("## Hôtes")
        lines.append("| IP | MAC | Vendeur | Hostname | Rôle |")
        lines.append("|---|---|---|---|---|")
        for h in hosts:
            lines.append(
                f"| {h.get('ip', '')} | {h.get('mac', '')} | "
                f"{h.get('vendor', '')} | {h.get('hostname', '')} | {h.get('role', '')} |"
            )
        lines.append("")

    # Exposures
    if exposures:
        lines.append("## Expositions")
        lines.append("| Hôte | Port | Service | Risque |")
        lines.append("|---|---|---|---|")
        for exp in exposures:
            lines.append(
                f"| {exp.get('host', exp.get('ip', ''))} | {exp.get('port', '')} | "
                f"{exp.get('service', '')} | {exp.get('risk', exp.get('severity', ''))} |"
            )
        lines.append("")

    # CVE
    if cves:
        lines.append("## CVE")
        lines.append("| CVE ID | Score | Sévérité | Description |")
        lines.append("|---|---|---|---|")
        for cve in cves:
            lines.append(
                f"| {cve.get('id', cve.get('cve_id', ''))} | {cve.get('score', '')} | "
                f"{cve.get('severity', '')} | {cve.get('description', '')[:80]} |"
            )
        lines.append("")

    lines.append("---")
    lines.append("*Generated by HARMATTAN*")
    return "\n".join(lines)
